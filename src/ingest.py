"""Load the raw export, validate schema and identities, report everything.

The two §3.2 identities are the data's own consistency proofs:
  1. VA/24 == VA Amount / Press hrs × 24   (rows with Press hrs > 0)
  2. mupnett == labmup + manadj            (what makes manadj a £ override)

Identity 1 failing means the file is not the schema we understand — raise.
Identity 2 failing is recorded on the ValidationReport; pricing.py checks
it and refuses to report rather than emit nonsense (§5.4), but the rest
of the system keeps working.

`#DIV/0!` trap: the export stores these as native Excel *error cells*
(data_type 'e'), not strings. pandas silently reads them as NaN, so they
are counted here with openpyxl before pandas ever sees the file.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd
import pandera.pandas as pa

SHEET = "Master Plain (Anon)"

IDENTITY_TOL = 1e-8  # §9: both identities hold to < 1e-8 max abs error

EXPECTED_COLUMNS = [
    "Title", "CustomerID", "Job Status", "SalesIn", "Year", "Month", "Week No",
    "SalesOut", "Quantity", "Sell Price", "Mup%", "VA Amount", "VA/24", "VA%",
    "VA/K", "Rebate", "Puchases", "Press hrs", "Impressions", "Handling",
    "Labour", "Paper", "labmup", "manadj", "mupnett", "Plates", "AmtInv",
    "Customer Name", "Rep", "Region", "Industry", "Work Type", "Product Type",
    "Binding Type", "Currency", "Ship date",
]


class IngestError(RuntimeError):
    """Raised when the file cannot be trusted as the known schema."""


@dataclass(frozen=True)
class ValidationReport:
    """Everything a human needs to trust (or distrust) an uploaded file."""

    source_name: str
    n_rows: int
    n_customers: int
    n_reps: int
    salesin_min: str  # ISO date
    salesin_max: str
    identity1_max_err: float
    identity2_max_err: float
    identity2_ok: bool  # pricing refuses to report when False
    va_pct_error_cells: int  # '#DIV/0!' stored as Excel error cells
    n_negative_sell_price: int  # quarantined later by clean
    n_null_salesout: int
    n_null_shipdate: int
    n_null_binding: int  # recoded to outsourced by clean, never imputed
    n_press_hrs_zero: int
    # undocumented nulls found in the real export — counted, handled, never silent
    n_null_manadj: int  # rows excluded from override analysis, identity 2 unchecked
    n_null_purchases: int
    n_null_product: int
    n_identity2_checked: int  # complete rows the identity was verified on
    schema_ok: bool


_schema = pa.DataFrameSchema(
    {
        "CustomerID": pa.Column(str),
        "Job Status": pa.Column(str),
        "SalesIn": pa.Column(pa.dtypes.Timestamp),
        "Year": pa.Column(int, pa.Check.in_range(2020, 2100)),
        "Month": pa.Column(int, pa.Check.in_range(1, 12)),
        "Quantity": pa.Column(float, coerce=True, nullable=True),
        "Sell Price": pa.Column(float, coerce=True),  # negatives are credits, kept
        "VA Amount": pa.Column(float, coerce=True),
        "VA/24": pa.Column(float, coerce=True),
        # misspelled in source, read as-is; 12 nulls in the real export (undocumented in §3.3)
        "Puchases": pa.Column(float, coerce=True, nullable=True),
        "Press hrs": pa.Column(float, pa.Check.ge(0), coerce=True),
        "labmup": pa.Column(float, coerce=True),
        # manadj/mupnett null x64 in the real export — identity 2 checked on complete rows;
        # pricing excludes those rows from override analysis and reports the count
        "manadj": pa.Column(float, coerce=True, nullable=True),
        "mupnett": pa.Column(float, coerce=True, nullable=True),
        "Rep": pa.Column(str),
        "Region": pa.Column(str),
        "Work Type": pa.Column(str),
        "Product Type": pa.Column(str, nullable=True),  # 1 null row in the real export
        "Binding Type": pa.Column(str, nullable=True),  # null = outsourced (data, not absence)
        "Currency": pa.Column(str, pa.Check.isin(["Stg", "Euro"])),
    },
    strict=False,  # extra columns tolerated; missing ones fail the explicit check below
)


def count_error_cells(path: Path, column: str, sheet: str = SHEET) -> int:
    """Count native Excel error cells (#DIV/0! etc.) in a column — pandas
    reads them as NaN, indistinguishable from blanks, so count first."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = header.index(column)
    n = sum(1 for row in ws.iter_rows(min_row=2) if row[idx].data_type == "e")
    wb.close()
    return n


def load_raw(path: Path) -> tuple[pd.DataFrame, ValidationReport]:
    """Read, schema-check and identity-check the export. Returns the raw
    frame (columns untouched, misspellings intact) plus the report."""
    df = pd.read_excel(path, sheet_name=SHEET)

    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        raise IngestError(f"{path.name}: missing expected columns {missing}")

    va_err_cells = count_error_cells(path, "VA%")

    try:
        _schema.validate(df, lazy=True)
        schema_ok = True
    except pa.errors.SchemaErrors as exc:
        raise IngestError(f"{path.name}: schema validation failed:\n{exc.failure_cases}") from exc

    # Identity 1 — VA/24 == VA Amount / Press hrs * 24, where Press hrs > 0
    hrs = df["Press hrs"].astype(float)
    with_hours = hrs > 0
    id1_err = (
        (df.loc[with_hours, "VA/24"] - df.loc[with_hours, "VA Amount"] / hrs[with_hours] * 24)
        .abs()
        .max()
    )
    if pd.isna(id1_err):
        id1_err = 0.0
    if id1_err >= IDENTITY_TOL:
        raise IngestError(
            f"{path.name}: identity VA/24 == VA Amount/Press hrs*24 broken "
            f"(max err {id1_err:.2e}) — file is not the understood schema"
        )

    # Identity 2 — mupnett == labmup + manadj, on complete rows (manadj is
    # null on some real rows). Recorded, not raised: pricing.py refuses to
    # report when this is False (§5.4).
    id2_complete = df[["mupnett", "labmup", "manadj"]].notna().all(axis=1)
    id2_err = float(
        (
            df.loc[id2_complete, "mupnett"]
            - (df.loc[id2_complete, "labmup"] + df.loc[id2_complete, "manadj"])
        )
        .abs()
        .max()
    )

    salesin = pd.to_datetime(df["SalesIn"])
    report = ValidationReport(
        source_name=path.name,
        n_rows=len(df),
        n_customers=int(df["CustomerID"].nunique()),
        n_reps=int(df["Rep"].nunique()),
        salesin_min=str(salesin.min().date()),
        salesin_max=str(salesin.max().date()),
        identity1_max_err=float(id1_err),
        identity2_max_err=id2_err,
        identity2_ok=bool(id2_err < IDENTITY_TOL),
        va_pct_error_cells=va_err_cells,
        n_negative_sell_price=int((df["Sell Price"] <= 0).sum()),
        n_null_salesout=int(df["SalesOut"].isna().sum()),
        n_null_shipdate=int(df["Ship date"].isna().sum()),
        n_null_binding=int(df["Binding Type"].isna().sum()),
        n_press_hrs_zero=int((hrs == 0).sum()),
        n_null_manadj=int(df["manadj"].isna().sum()),
        n_null_purchases=int(df["Puchases"].isna().sum()),
        n_null_product=int(df["Product Type"].isna().sum()),
        n_identity2_checked=int(id2_complete.sum()),
        schema_ok=schema_ok,
    )
    return df, report
